#Run streamlit environemnt locally: https://docs.streamlit.io/library/get-started/installation
#df documentation: https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.html
# Import the libraries BeautifulSoup and os
import streamlit as st
from bs4 import BeautifulSoup as bs
import requests
from datetime import datetime
import time
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell
from io import StringIO


#Cache it so that it doesn't request the url again and again and mess up storage
@st.cache_resource
def request_url(Link):
    try:
        page_to_scrape = requests.get(Link.strip())
    except:
        st.write("Unable to retrieve the next link: " + Link.strip())
        st.write("There is possibly an issue with the csv file/link you uploaded")
    soup = bs(page_to_scrape.text, "html.parser")
    return soup


def print_payment_options(categoryDescDiv, categoryName, numChosenCategories):
    sheet=st.session_state.sheet
    workbook=st.session_state.workbook
    paymentcontainer = st.empty()
    containerToWriteTo=paymentcontainer.container()

    #for writing to file
    ascii_value_of_current_category=ord('F')+numChosenCategories-1
    #should if you have uploaded 2 or more film festival links- not coded though
    row_to_write_to=st.session_state.nextLineToWriteOn+1

    all_profile_deadlines=categoryDescDiv.findChild('ul' , recursive=False)
    all_li = all_profile_deadlines.findChildren()
    paymentOptions=0
    dictOfButtons={}
    ReturnValue=""
    paymentOptionChosen=""
    numDeadlines=0
    numButtons=1
    for index, current_li in enumerate(all_li):
        with containerToWriteTo:
            #if it has not already passed
            if 'is-outdated' not in current_li['class']:
                #Only consider deadline and payment option
                if 'CategoryName' in current_li['class'] or 'Fee' in current_li['class']:
                    #If the current element is a deadline label/name
                    if 'CategoryName' in current_li['class']:
                        currentDeadlineName = current_li.text.strip()
                        numDeadlines=numDeadlines+1
                        st.markdown('<h5 style="color:purple"> '+ str(numDeadlines)+ '. "'+ currentDeadlineName + '" deadline </h5>', unsafe_allow_html=True)
                        st.markdown('<h6 style="color:violet"> For submitting your film to the ' + categoryName + ' category by the <span style="color:white">'+ currentDeadlineName + ' </span>deadline, here are the payment options </h6>', unsafe_allow_html=True)
                    #If the current element is a payment option, print it out and after printing it out, ask the user which payment option they want
                    elif 'Fee' in current_li['class']:
                        paymentOptions=paymentOptions+1
                        currOption=current_li.text.replace("\n", " ")
                        dictOfButtons[currOption] = st.button(f''':violet[{numButtons}. {currOption}\n{categoryName}]''')
                        numButtons=numButtons+1
                        #if the next element is a category element, then that means there are no more payment options 
                        next = index + 1
                        if next <= len(all_li) :
                            #avoid index out of bounds error at the end
                            if (next==len(all_li) or 'Fee' not in all_li[next]['class']):
                                #if True is a dictionary value, that means that a button has been pressed
                                while True not in dictOfButtons.values():
                                    time.sleep(0.5)
                                #identify which key has the value of true. This key is the key that was selected and represents the user's payment choice
                                for key, value in dictOfButtons.items():
                                    if value == True:
                                        #key is the chosen payment option
                                        ReturnValue+=key + ','

                                        #write to workbook
                                        #add payment option to the workbook and then move onto the next row
                                        currRowPayment=str(row_to_write_to)
                                        #worksheet.write(chr(ascii_value_of_current_category)+currRowPayment, key)
                                        st.session_state.sheet[chr(ascii_value_of_current_category)+currRowPayment]=key
                                        row_to_write_to=row_to_write_to+1
                                        
                                        paymentOptions=0
                                        dictOfButtons={}
                                        
                                        #clear the paymentcontainer and create a new one to write to
                                        paymentcontainer.empty()
                                        containerToWriteTo=paymentcontainer.container()
                                        #continue
                                        break
    paymentcontainer.empty()
    return ReturnValue
            

                                   
                                

def ask_about_categories(soup, ascii_value_of_col_of_first_category,value_of_row_to_write_to):
    workbook=st.session_state.workbook
    sheet=st.session_state.sheet
    DivOfCategories=soup.find('ul', attrs={'class':'ProfileCategories'})
    #grab the first <li data-ga-category= Festival Show - Sidebar>, which is the div that contains the information for the first category
    currCategoryDiv= DivOfCategories.findChild("li", recursive=False)
    currNumOfCategoriesChosen=0
    totalNumOfCategories=0
    while currCategoryDiv!=None:
        #Deal with category name
        #(1)grab first child div which should be <div class="profile-categories__title">
        #(2)grab the first span which has the class festival-category-name which should contain the name of the category
        container=st.empty()
        with container.container():
        
            name=currCategoryDiv.findAll('div', attrs={'class':'profile-categories__title'}, recursive=False)[0].findAll('span', attrs={'class':'festival-category-name'}, recursive=False)[0]

            #Deal with category description
            #find the first occurence of <div class="profile-categories_content">
            currCategoryDescriptionDiv=currCategoryDiv.findAll('div', attrs={'class':'profile-categories__content'}, recursive=False)[0]
            description=currCategoryDescriptionDiv.findChild('div' , recursive=False)
            totalNumOfCategories=totalNumOfCategories+1
            st.markdown('<h3> Category '+ str(totalNumOfCategories)+ ': ' + name.text.strip() + '</h3> \n' + '<h5>' + description.text.strip() + '</h5>', unsafe_allow_html=True)
            
            
            st.write("Do you want to apply to this category?")
            col1, col2 = st.columns(2)
            with col1:
                yes_button=st.button(f''':green[Yes, apply to {name.text.strip()}]''')
            with col2:
                no_button=st.button(f''':red[No, do not apply to {name.text.strip()}]''')
            while not yes_button and not no_button:
                time.sleep(0.5)

        if yes_button or no_button:
            container.empty()
            if yes_button:
                #st.info(f''':green[You are applying to the category: {name.text}]''')

                #write to the workbook
                #add category name into the workbook
                currentCell= chr(ascii_value_of_col_of_first_category+currNumOfCategoriesChosen)+str(st.session_state.nextLineToWriteOn)
                #worksheet.write(currentCell, name.text)
                st.session_state.sheet[currentCell]=name.text
                #move to the next column
                currNumOfCategoriesChosen=currNumOfCategoriesChosen+1

                paymentChoices= print_payment_options(currCategoryDescriptionDiv, name.text.strip(), currNumOfCategoriesChosen)
                st.info(f''':violet[You chose payment option(s) '{paymentChoices}' for all the deadlines in] :green['{name.text}'] :violet[category]''') 
            #else:
                #st.info(f''':red[You are NOT applying to the category: {name.text}]''')

        #move onto next category
        currCategoryDiv= currCategoryDiv.find_next_sibling("li")
        #break

    if currNumOfCategoriesChosen >= st.session_state.maxNumOfCategories:
       st.session_state.maxNumOfCategories=currNumOfCategoriesChosen   
    return currNumOfCategoriesChosen

def changeSubmissionState(nameOfFestSubmitted, maincontainer):
    st.session_state.submitted = nameOfFestSubmitted
    st.session_state.downloaded = True
    maincontainer.empty()

def runFilmFestivals(link):
    sheet=st.session_state.sheet
    workbook=st.session_state.workbook
    maincontainer=st.empty()
    with maincontainer.container():
        #htmlOfLink='Slamdance.html'
        soup=request_url(link)
        present = datetime.now()

        NameOfFestival= soup.find('a', attrs={"data-ga-label":"Festival Name"})
        st.markdown('<h1 style="color:blue;">' + NameOfFestival.text.strip() + '</h1>', unsafe_allow_html=True)



        #the column where the first category will go is column F
        ascii_value_of_col_of_first_category=ord('F')
        #we just wrote to the header, so now we want to write to the next row
        value_of_row_to_write_to=st.session_state.nextLineToWriteOn

        numOfCategoriesChosen=ask_about_categories(soup,ascii_value_of_col_of_first_category,value_of_row_to_write_to)

        #move onto the next line after having printed all category names
        value_of_row_to_write_to= st.session_state.nextLineToWriteOn +1

        #Find all the deadline timings
        Deadlines = soup.findAll('time', attrs={'class':'ProfileFestival-datesDeadlines-time'})
        #Find all the deadline names
        DeadlineLabels = soup.findAll('div', attrs={"class":"ProfileFestival-datesDeadlines-deadline"})

        for deadline, deadlinelabel in zip(Deadlines, DeadlineLabels):
            if(deadlinelabel.text.strip() == "Event Date"):
                eventdate=deadline.text.strip()


        #value_of_row_to_write_to=value_of_row_to_write_to+1
        numOfDeadlinesPrinted=0
        for deadline, deadlinelabel in zip(Deadlines, DeadlineLabels):
            #the event date does not follow strptime properties
            #notification date is unnecessary
            if(deadlinelabel.text.strip() != "Event Date" and deadlinelabel.text.strip() != "Notification Date"):
                #convert the date from text into datetime format
                currentdate= datetime.strptime(deadline.text.strip(), "%B %d, %Y")
                #if the deadline has not already passed then it is added to 
                # the file
                if (currentdate>present):
                    #write to the worksheet
                    currRow=str(value_of_row_to_write_to)
                    #worksheet.write('D'+currRow, deadlinelabel.text.strip())
                    #worksheet.write('E'+currRow, deadline.text.strip())
                    st.session_state.sheet['D'+currRow]=deadlinelabel.text.strip()
                    st.session_state.sheet['E'+currRow]=deadline.text.strip()
                    value_of_row_to_write_to= value_of_row_to_write_to +1
                    #increment number of deadlines printed
                    numOfDeadlinesPrinted=numOfDeadlinesPrinted + 1

        #write to workbook
        #merge cells for currFestName, currFestDate, and currFestLink
       
        value_of_row_to_write_to=st.session_state.nextLineToWriteOn+1
        endingRow=value_of_row_to_write_to + numOfDeadlinesPrinted -1
        if endingRow > value_of_row_to_write_to:
            #worksheet.merge_range('A'+ str(value_of_row_to_write_to)+ '\:A'+ str(endingRow),    NameOfFestival.text.strip(), merge_format)
            #worksheet.merge_range('B'+ str(value_of_row_to_write_to)+ '\:B'+ str(endingRow),    eventdate, merge_format)
            #worksheet.merge_range('C'+ str(value_of_row_to_write_to)+ '\:C'+ str(endingRow),  link , merge_format)
            st.session_state.sheet.merge_cells(start_row=value_of_row_to_write_to, start_column=1, end_row=endingRow, end_column=1)
            st.session_state.sheet.merge_cells(start_row=value_of_row_to_write_to, start_column=2, end_row=endingRow, end_column=2)
            st.session_state.sheet.merge_cells(start_row=value_of_row_to_write_to, start_column=3, end_row=endingRow, end_column=3)
        #if there is only one deadline, then no need to merge any rows
        #AND after merging rows, write information to it
        st.session_state.sheet['A'+str(value_of_row_to_write_to)] = NameOfFestival.text.strip()
        st.session_state.sheet['B'+str(value_of_row_to_write_to)]=eventdate
        st.session_state.sheet['C'+str(value_of_row_to_write_to)]=link

        #the number of columns is Festival Name, Deadline Label, Deadline, and all categories. if no categories were chosen, then only 6 columns should be present
        if st.session_state.maxNumOfCategories>0:
            numOfColumns= 6 + (st.session_state.maxNumOfCategories-1)
        else:
            numOfColumns= 6
        
        
        rangeOfWorksheet='A1:' + chr(ord('@')+numOfColumns)+ str(numOfDeadlinesPrinted+2)
        #change column width
        st.session_state.sheet.column_dimensions['A'].width = 17
        st.session_state.sheet.column_dimensions['B'].width = 17
        st.session_state.sheet.column_dimensions['C'].width = 23
        st.session_state.sheet.column_dimensions['D'].width = 18
        st.session_state.sheet.column_dimensions['E'].width = 16
        #change column width of all categories
        i=0
        while i<st.session_state.maxNumOfCategories:
            currentColToChange=chr(ord('F')+i)
            #print("current col to change: " + str(currentColToChange))
            st.session_state.sheet.column_dimensions[currentColToChange].width = 16
            i=i+1

        # set header format to gray background, increase font size and increase bottom border
        grayFill = PatternFill(start_color='D0CECE', end_color='D0CECE',
                       fill_type='solid')
        fontStyle = Font(size = "14", bold=True)
        for cell in st.session_state.sheet[1:1]:
            cell.fill = grayFill
            cell.font = fontStyle

        #align all cells center
        for col in st.session_state.sheet.columns:
            for cell in col:
                # openpyxl styles aren't mutable,
                # so you have to create a copy of the style, modify the copy, then set it back
                if cell.value is None:
                    blackFill = PatternFill("solid", fgColor="000000")
                    cell.fill=blackFill
                else:
                    alignment_obj = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = alignment_obj
        #both row 1 and column A frozen
        st.session_state.sheet.freeze_panes = "B2"

        st.session_state.workbook.save("ExcelFile.xlsx")
    
    dowloadbtncontainer=st.empty()
    with dowloadbtncontainer:
        _="""Download the formatted excel sheet"""
        st.write("Done with collecting data. Please wait a couple minutes as your data gets downloaded")
        with open("ExcelFile.xlsx", "rb") as template_file:
            template_byte = template_file.read()
        #adding download button to dowload excel file
        st.download_button(
            label='📥 Download Result(s) as Excel Sheet', 
            data=template_byte ,
            file_name= 'Excel_Film_Festivals.xlsx',
            on_click=changeSubmissionState(link, maincontainer)
        )
    st.session_state.downloadbtn=dowloadbtncontainer
    #number of deadlines + categories + next line to write on
    st.session_state.nextLineToWriteOn= st.session_state.nextLineToWriteOn+ numOfDeadlinesPrinted+1+1

    


def uploadcsv():
    placeholder=st.empty()
    with placeholder.container():
        uploaded_file = placeholder.file_uploader("Choose a csv file that has filmfreeway links to film festivals. The delimiter MUST be a newline character", type={"csv", "txt"})
    if uploaded_file is not None:
        placeholder.empty()
        # To read file as string:
        stringio = StringIO(uploaded_file.getvalue().decode("utf-8"))
        string_data = stringio.read()
        #st.write(string_data)
        festival_links= string_data.split("\n")
        return festival_links
    else:
        return None

def insertsinglelink():
    placeholder2=st.empty()
    with placeholder2.container():
        Festlink = st.text_input('Please enter ONLY one filmfreeway url')
        submit_button = st.button("Submit Link")
    if submit_button: 
        if Festlink is not None:
            placeholder2.empty()
            return Festlink
        

def choose_num_of_festival_links():
    festival_links=[]
    placeholder = st.empty()
    with placeholder.container():
        #st.write(str(categoryapplicationuniquekey))
        st.write("How many FilmFreeway links do you want to submit?")
        col1, col2 = st.columns(2)
        with col1:
            one_button=st.button("One")
        with col2:
            more_button=st.button("More Than One (upload CSV file)")
        while not one_button and not more_button:
            time.sleep(0.5)
    if one_button or more_button:
        placeholder.empty()
        if one_button:
            singleLink=insertsinglelink()
            while singleLink is None:
                time.sleep(0.5)
            #once single link has a vlue in it then append it to the end of festival_links
            festival_links.append(singleLink)
            return festival_links
        else:
            festival_links=uploadcsv()
            while festival_links is None:
                time.sleep(0.5)
            if festival_links is not None:
                return festival_links
        


#executed the first time the app is run
if 'submitted' not in st.session_state:
    #maxNumOfCategoriesChosen=0
    Links=choose_num_of_festival_links()
    #st.write(Links)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1']='Festival Name'
    sheet['B1']='Festival Date'
    sheet['C1']='Submission Link'
    sheet['D1']='Deadline Label'
    sheet['E1']='Deadline'
    sheet['F1']='Categories'
    st.session_state.Links=Links
    st.session_state.workbook=wb
    st.session_state.sheet=sheet

    st.session_state.maxNumOfCategories=0
    st.session_state.nextLineToWriteOn=2
    st.session_state.linkIdx=0
    #save links, workbook, and worksheet as session state so we can use later on
    runFilmFestivals(Links[st.session_state.linkIdx])
    st.session_state.linkIdx=st.session_state.linkIdx+1
    #add 1 because header was added here
    #st.session_state.nextLineToWriteOn= st.session_state.nextLineToWriteOn+1
    

#after you have downloaded the file
else:
    #traverse Links until you reach the link that was just equal to it. Then, if there is a next one,move onto the next one. else, tell the user you have printed out everything
    #justDownloaded=st.session_state.submitted 
    #st.info("You have downloaded info for " + justDownloaded)
    #once you click downloaded, the session state gets updated and you delete the button then u set it to undownloaded for the next festival
    if st.session_state.downloaded ==True:
        st.session_state.downloaded =False
        st.session_state.downloadbtn.empty()
    Links=st.session_state.Links
    print(st.session_state.linkIdx)
    if (st.session_state.linkIdx>= len(Links)):
        st.write("Thank you for using this website. We are done with the link(s)")
        #st.write(st.session_state.linkIdx)
        st.cache_resource.clear()
    else:
        runFilmFestivals(Links[st.session_state.linkIdx])
        st.session_state.linkIdx=st.session_state.linkIdx+1


















